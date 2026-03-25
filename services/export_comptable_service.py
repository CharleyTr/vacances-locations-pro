"""
Service d'export comptable Excel — génère un classeur par année/propriété.
"""
import io
from datetime import date
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# ── Couleurs ─────────────────────────────────────────────────────────────────
C_HEADER_BG  = "1565C0"   # bleu foncé
C_HEADER_FG  = "FFFFFF"   # blanc
C_TOTAL_BG   = "E3F2FD"   # bleu clair
C_ALT_BG     = "F7FBFF"   # bleu très clair (lignes paires)
C_TITLE_BG   = "0D47A1"   # bleu très foncé
C_GREEN      = "2E7D32"
C_RED        = "C62828"

EUR   = '#,##0.00 €'
EUR0  = '#,##0 €'
PCT   = '0.0%'
DATE_FMT = 'DD/MM/YYYY'

def _fill(hex_color):
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)

def _font(bold=False, color="000000", size=10):
    return Font(bold=bold, color=color, size=size, name="Arial")

def _border():
    thin = Side(style="thin", color="BDBDBD")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _right():
    return Alignment(horizontal="right", vertical="center")

def _col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width


# ─────────────────────────────────────────────────────────────────────────────
# FEUILLE 1 : Détail des réservations
# ─────────────────────────────────────────────────────────────────────────────

def _sheet_detail(wb, df, annee, prop_nom):
    ws = wb.active
    ws.title = "Réservations"
    ws.freeze_panes = "A3"

    # Titre
    ws.merge_cells("A1:N1")
    ws["A1"] = f"EXPORT COMPTABLE — {prop_nom.upper()} — {annee}"
    ws["A1"].font    = _font(bold=True, color=C_HEADER_FG, size=12)
    ws["A1"].fill   = _fill(C_TITLE_BG)
    ws["A1"].alignment = _center()
    ws.row_dimensions[1].height = 28

    # En-têtes
    COLS = [
        ("N° Réservation", 20), ("Plateforme", 14), ("Client", 22),
        ("Arrivée", 12),        ("Départ", 12),      ("Nuits", 7),
        ("CA Brut", 12),        ("Commission", 12),  ("Ménage", 10),
        ("Taxes séjour", 12),   ("Frais CB", 11),    ("CA Net", 12),
        ("Payé", 8),            ("Observation", 22),
    ]
    for c, (title, width) in enumerate(COLS, 1):
        cell = ws.cell(row=2, column=c, value=title)
        cell.font      = _font(bold=True, color=C_HEADER_FG, size=9)
        cell.fill      = _fill(C_HEADER_BG)
        cell.alignment = _center()
        cell.border    = _border()
        _col_width(ws, c, width)
    ws.row_dimensions[2].height = 22

    # Données
    df_s = df[df["plateforme"] != "Fermeture"].copy()
    df_s = df_s.sort_values("date_arrivee")

    def _fv(row, col, default=0):
        v = row.get(col, default)
        return float(v) if pd.notna(v) and v != "" else default

    for i, (_, row) in enumerate(df_s.iterrows(), 3):
        bg = C_ALT_BG if i % 2 == 0 else "FFFFFF"
        vals = [
            row.get("numero_reservation", ""),
            row.get("plateforme", ""),
            row.get("nom_client", ""),
            row.get("date_arrivee"),
            row.get("date_depart"),
            int(_fv(row, "nuitees")),
            _fv(row, "prix_brut"),
            _fv(row, "commission"),
            _fv(row, "prix_menage"),
            _fv(row, "taxes_sejour"),
            _fv(row, "frais_cb"),
            _fv(row, "prix_net"),
            "✓" if row.get("paye") else "—",
            "",
        ]
        fmts = [None, None, None, DATE_FMT, DATE_FMT, "0",
                EUR0, EUR0, EUR0, EUR0, EUR0, EUR0, None, None]

        for c, (val, fmt) in enumerate(zip(vals, fmts), 1):
            cell = ws.cell(row=i, column=c, value=val)
            cell.font      = _font(size=9)
            cell.fill      = _fill(bg)
            cell.border    = _border()
            if fmt:
                cell.number_format = fmt
            if c >= 7 and c <= 12:
                cell.alignment = _right()
            # Couleur colonne CA Net
            if c == 12:
                cell.font = _font(bold=True, color=C_GREEN, size=9)
            if c == 13 and val == "—":
                cell.font = _font(color=C_RED, size=9)

    # Ligne TOTAL
    n = len(df_s) + 3  # ligne total
    ws.cell(row=n, column=1, value="TOTAL").font = _font(bold=True, size=9)
    ws.cell(row=n, column=1).fill = _fill(C_TOTAL_BG)

    total_cols = {7: "prix_brut", 8: "commission", 9: "prix_menage",
                  10: "taxes_sejour", 11: "frais_cb", 12: "prix_net", 6: "nuitees"}
    for col_idx, field in total_cols.items():
        first_row = 3
        last_row  = n - 1
        col_letter = get_column_letter(col_idx)
        cell = ws.cell(row=n, column=col_idx,
                       value=f"=SUM({col_letter}{first_row}:{col_letter}{last_row})")
        cell.font   = _font(bold=True, color=C_TITLE_BG, size=9)
        cell.fill   = _fill(C_TOTAL_BG)
        cell.border = _border()
        cell.number_format = "0" if col_idx == 6 else EUR0
        cell.alignment = _right()

    ws.row_dimensions[n].height = 18


# ─────────────────────────────────────────────────────────────────────────────
# FEUILLE 2 : Récapitulatif mensuel
# ─────────────────────────────────────────────────────────────────────────────

def _sheet_mensuel(wb, df, annee):
    ws = wb.create_sheet("Mensuel")
    ws.freeze_panes = "A3"

    MOIS_FR = ["Janvier","Février","Mars","Avril","Mai","Juin",
               "Juillet","Août","Septembre","Octobre","Novembre","Décembre"]

    # Titre
    ws.merge_cells("A1:H1")
    ws["A1"] = f"RÉCAPITULATIF MENSUEL {annee}"
    ws["A1"].font = _font(bold=True, color=C_HEADER_FG, size=11)
    ws["A1"].fill = _fill(C_TITLE_BG)
    ws["A1"].alignment = _center()
    ws.row_dimensions[1].height = 26

    headers = ["Mois", "Réservations", "Nuits", "CA Brut", "Commissions", "Ménages", "CA Net", "Taux Occ."]
    widths  = [14, 14, 8, 14, 14, 12, 14, 12]
    for c, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.font = _font(bold=True, color=C_HEADER_FG, size=9)
        cell.fill = _fill(C_HEADER_BG)
        cell.alignment = _center()
        cell.border = _border()
        _col_width(ws, c, w)

    df_reel = df[df["plateforme"] != "Fermeture"].copy()
    df_reel["mois"] = pd.to_datetime(df_reel["date_arrivee"]).dt.month

    import calendar
    for i, (mois_num, nom_mois) in enumerate(zip(range(1, 13), MOIS_FR), 3):
        df_m = df_reel[df_reel["mois"] == mois_num]
        nb_jours = calendar.monthrange(int(annee), mois_num)[1]
        nuits = float(df_m["nuitees"].fillna(0).sum()) if "nuitees" in df_m.columns else 0
        taux  = nuits / nb_jours if nb_jours > 0 else 0

        bg = C_ALT_BG if i % 2 == 0 else "FFFFFF"
        vals = [
            nom_mois,
            len(df_m),
            int(nuits),
            float(df_m["prix_brut"].fillna(0).sum())     if "prix_brut"    in df_m.columns else 0,
            float(df_m["commission"].fillna(0).sum())    if "commission"   in df_m.columns else 0,
            float(df_m["prix_menage"].fillna(0).sum())   if "prix_menage"  in df_m.columns else 0,
            float(df_m["frais_cb"].fillna(0).sum())      if "frais_cb"     in df_m.columns else 0,
            float(df_m["prix_net"].fillna(0).sum())      if "prix_net"     in df_m.columns else 0,
            taux,
        ]
        fmts = [None, "0", "0", EUR0, EUR0, EUR0, EUR0, PCT]
        for c, (val, fmt) in enumerate(zip(vals, fmts), 1):
            cell = ws.cell(row=i, column=c, value=val)
            cell.font   = _font(size=9)
            cell.fill   = _fill(bg)
            cell.border = _border()
            if fmt: cell.number_format = fmt
            if c >= 3: cell.alignment = _right()
            if c == 7: cell.font = _font(bold=True, color=C_GREEN, size=9)

    # Total annuel
    tot_row = 15
    ws.cell(row=tot_row, column=1, value="TOTAL ANNUEL").font = _font(bold=True, size=9)
    ws.cell(row=tot_row, column=1).fill = _fill(C_TOTAL_BG)
    for col_idx, fmt in [(2,"0"),(3,"0"),(4,EUR0),(5,EUR0),(6,EUR0),(7,EUR0)]:
        col_letter = get_column_letter(col_idx)
        cell = ws.cell(row=tot_row, column=col_idx,
                       value=f"=SUM({col_letter}3:{col_letter}14)")
        cell.font   = _font(bold=True, color=C_TITLE_BG, size=9)
        cell.fill   = _fill(C_TOTAL_BG)
        cell.border = _border()
        cell.number_format = fmt
        cell.alignment = _right()
    # Taux moyen
    cell = ws.cell(row=tot_row, column=8, value=f"=AVERAGE(H3:H14)")
    cell.font = _font(bold=True, color=C_TITLE_BG, size=9)
    cell.fill = _fill(C_TOTAL_BG)
    cell.border = _border()
    cell.number_format = PCT
    cell.alignment = _right()


# ─────────────────────────────────────────────────────────────────────────────
# FEUILLE 3 : Récap par plateforme
# ─────────────────────────────────────────────────────────────────────────────

def _sheet_plateforme(wb, df):
    ws = wb.create_sheet("Par plateforme")

    ws.merge_cells("A1:G1")
    ws["A1"] = "RÉCAPITULATIF PAR PLATEFORME"
    ws["A1"].font = _font(bold=True, color=C_HEADER_FG, size=11)
    ws["A1"].fill = _fill(C_TITLE_BG)
    ws["A1"].alignment = _center()
    ws.row_dimensions[1].height = 26

    headers = ["Plateforme", "Réservations", "Nuits", "CA Brut", "Commissions", "CA Net", "% CA Net"]
    widths  = [18, 14, 10, 14, 14, 14, 12]
    for c, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.font = _font(bold=True, color=C_HEADER_FG, size=9)
        cell.fill = _fill(C_HEADER_BG)
        cell.alignment = _center()
        cell.border = _border()
        _col_width(ws, c, w)

    df_reel = df[df["plateforme"] != "Fermeture"].copy()
    plateformes = sorted(df_reel["plateforme"].dropna().unique())
    ca_net_total = float(df_reel["prix_net"].fillna(0).sum()) if "prix_net" in df_reel.columns else 1

    for i, plat in enumerate(plateformes, 3):
        df_p = df_reel[df_reel["plateforme"] == plat]
        nuits = float(df_p["nuitees"].fillna(0).sum()) if "nuitees" in df_p.columns else 0
        ca_brut = float(df_p["prix_brut"].fillna(0).sum()) if "prix_brut" in df_p.columns else 0
        comm    = float(df_p["commission"].fillna(0).sum()) if "commission" in df_p.columns else 0
        ca_net  = float(df_p["prix_net"].fillna(0).sum()) if "prix_net" in df_p.columns else 0
        pct     = ca_net / ca_net_total if ca_net_total > 0 else 0
        bg = C_ALT_BG if i % 2 == 0 else "FFFFFF"
        for c, (val, fmt) in enumerate(zip(
            [plat, len(df_p), int(nuits), ca_brut, comm, ca_net, pct],
            [None, "0", "0", EUR0, EUR0, EUR0, PCT]
        ), 1):
            cell = ws.cell(row=i, column=c, value=val)
            cell.font = _font(size=9)
            cell.fill = _fill(bg)
            cell.border = _border()
            if fmt: cell.number_format = fmt
            if c >= 2: cell.alignment = _right()
            if c == 6: cell.font = _font(bold=True, color=C_GREEN, size=9)

    # Total
    tot = len(plateformes) + 3
    ws.cell(row=tot, column=1, value="TOTAL").font = _font(bold=True, size=9)
    ws.cell(row=tot, column=1).fill = _fill(C_TOTAL_BG)
    for col_idx, fmt in [(2,"0"),(3,"0"),(4,EUR0),(5,EUR0),(6,EUR0)]:
        col_letter = get_column_letter(col_idx)
        cell = ws.cell(row=tot, column=col_idx,
                       value=f"=SUM({col_letter}3:{col_letter}{tot-1})")
        cell.font = _font(bold=True, color=C_TITLE_BG, size=9)
        cell.fill = _fill(C_TOTAL_BG)
        cell.border = _border()
        cell.number_format = fmt
        cell.alignment = _right()


# ─────────────────────────────────────────────────────────────────────────────
# FEUILLE 4 : KPIs synthèse
# ─────────────────────────────────────────────────────────────────────────────

def _sheet_kpis(wb, df, annee, prop_nom):
    ws = wb.create_sheet("Synthèse")
    _col_width(ws, 1, 30)
    _col_width(ws, 2, 20)

    ws.merge_cells("A1:B1")
    ws["A1"] = f"SYNTHÈSE — {prop_nom.upper()} — {annee}"
    ws["A1"].font = _font(bold=True, color=C_HEADER_FG, size=12)
    ws["A1"].fill = _fill(C_TITLE_BG)
    ws["A1"].alignment = _center()
    ws.row_dimensions[1].height = 30

    df_reel = df[df["plateforme"] != "Fermeture"].copy()
    import calendar
    nb_jours = 366 if calendar.isleap(int(annee)) else 365
    nuits     = float(df_reel["nuitees"].fillna(0).sum()) if "nuitees" in df_reel.columns else 0
    ca_brut   = float(df_reel["prix_brut"].fillna(0).sum()) if "prix_brut" in df_reel.columns else 0
    comm      = float(df_reel["commission"].fillna(0).sum()) if "commission" in df_reel.columns else 0
    menage    = float(df_reel["prix_menage"].fillna(0).sum()) if "prix_menage" in df_reel.columns else 0
    taxes     = float(df_reel["taxes_sejour"].fillna(0).sum()) if "taxes_sejour" in df_reel.columns else 0
    ca_net    = float(df_reel["prix_net"].fillna(0).sum()) if "prix_net" in df_reel.columns else 0
    nb_res    = len(df_reel)
    taux_occ  = nuits / nb_jours if nb_jours > 0 else 0
    rev_nuit  = ca_net / nuits if nuits > 0 else 0
    duree_moy = nuits / nb_res if nb_res > 0 else 0

    kpis = [
        ("", "", None),
        ("📊 ACTIVITÉ", "", None),
        ("Nombre de réservations",   nb_res,                       "0"),
        ("Nuits louées",             int(nuits),                   "0"),
        ("Durée moyenne séjour",     round(duree_moy, 1),          "0.0"),
        ("Taux d'occupation",        taux_occ,                     PCT),
        ("", "", None),
        ("💶 FINANCIER", "", None),
        ("CA Brut total",            ca_brut,                      EUR0),
        ("Commissions plateformes",  comm,                         EUR0),
        ("Frais ménage",             menage,                       EUR0),
        ("Frais CB",                 frais_cb,                     EUR0),
        ("Taxes de séjour",          taxes,                        EUR0),
        ("CA Net total",             ca_net,                       EUR0),
        ("Revenu moyen / nuit",      rev_nuit,                     EUR0),
        ("", "", None),
        ("📋 INFORMATIONS", "", None),
        ("Propriété",                prop_nom,                     None),
        ("Année fiscale",            str(annee),                   None),
        ("Date d'export",            date.today().strftime("%d/%m/%Y"), None),
    ]

    for r, (label, val, fmt) in enumerate(kpis, 2):
        is_section = label.startswith("📊") or label.startswith("💶") or label.startswith("📋")
        is_empty   = label == ""

        cell_l = ws.cell(row=r, column=1, value=label)
        cell_v = ws.cell(row=r, column=2, value=val if not is_section else "")

        if is_section:
            cell_l.font = _font(bold=True, color=C_HEADER_FG, size=10)
            cell_l.fill = _fill(C_HEADER_BG)
            cell_v.fill = _fill(C_HEADER_BG)
            ws.merge_cells(f"A{r}:B{r}")
            ws.row_dimensions[r].height = 20
        elif not is_empty:
            cell_l.font = _font(size=10)
            cell_v.font = _font(bold=True, size=10,
                                color=C_GREEN if label in ("CA Net total","Revenu moyen / nuit") else "000000")
            cell_l.border = cell_v.border = _border()
            cell_v.alignment = _right()
            if fmt: cell_v.number_format = fmt


# ─────────────────────────────────────────────────────────────────────────────
# EXPORT PRINCIPAL
# ─────────────────────────────────────────────────────────────────────────────

def generate_export(df: pd.DataFrame, annee: int, prop_nom: str) -> bytes:
    """
    Génère le classeur Excel et retourne les bytes pour st.download_button.
    """
    df_an = df[df["annee"] == annee].copy() if "annee" in df.columns else df.copy()

    # Normaliser les noms de colonnes (différents selon source Booking/Airbnb)
    if "commissions" in df_an.columns and "commission" not in df_an.columns:
        df_an["commission"] = df_an["commissions"]
    if "menage" in df_an.columns and "prix_menage" not in df_an.columns:
        df_an["prix_menage"] = df_an["menage"]
    if "frais_menage" in df_an.columns and "prix_menage" not in df_an.columns:
        df_an["prix_menage"] = df_an["frais_menage"]
    if "taxes_sejour" not in df_an.columns:
        df_an["taxes_sejour"] = 0.0

    wb = Workbook()
    _sheet_detail(wb, df_an, annee, prop_nom)
    _sheet_mensuel(wb, df_an, annee)
    _sheet_plateforme(wb, df_an)
    _sheet_kpis(wb, df_an, annee, prop_nom)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
