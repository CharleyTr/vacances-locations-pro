"""
Service de génération de rapports Excel.
Produit un fichier .xlsx avec 4 onglets :
  - Résumé annuel
  - Détail mensuel
  - Liste réservations
  - Prévisions
"""
import io
import pandas as pd
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              numbers as xl_numbers)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import SeriesLabel

# ── Palette couleurs ──────────────────────────────────────────────────────────
BLEU_DARK  = "1A237E"
BLEU_MED   = "3949AB"
BLEU_LIGHT = "C5CAE9"
GRIS_FOND  = "F5F5F5"
VERT       = "2E7D32"
ROUGE      = "C62828"
ORANGE     = "E65100"
BLANC      = "FFFFFF"

_thin = Side(style="thin", color="BDBDBD")
_border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _header_cell(ws, row, col, value, bg=BLEU_DARK, fg=BLANC, bold=True, size=11):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, color=fg, size=size, name="Arial")
    c.fill = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = _border
    return c


def _data_cell(ws, row, col, value, fmt=None, bold=False, bg=None, align="right"):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Arial", size=10, bold=bold)
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border = _border
    if fmt:
        c.number_format = fmt
    if bg:
        c.fill = PatternFill("solid", start_color=bg)
    return c


def generate_report(df: pd.DataFrame, prop_nom: str, annee: int) -> bytes:
    """
    Génère le rapport Excel et retourne les bytes du fichier.
    """
    wb = Workbook()
    wb.remove(wb.active)  # Supprimer la feuille par défaut

    df = df.copy()
    df["date_arrivee"] = pd.to_datetime(df["date_arrivee"])
    df["date_depart"]  = pd.to_datetime(df["date_depart"])

    # Exclure les fermetures des calculs financiers
    df_reel = df[df["plateforme"] != "Fermeture"].copy()

    _sheet_resume(wb, df_reel, prop_nom, annee)
    _sheet_mensuel(wb, df_reel, annee)
    _sheet_reservations(wb, df_reel)
    _sheet_previsions(wb, df, prop_nom)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Onglet 1 : Résumé ─────────────────────────────────────────────────────────

def _sheet_resume(wb, df, prop_nom, annee):
    ws = wb.create_sheet("📊 Résumé")
    ws.sheet_view.showGridLines = False

    # Titre
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = f"BILAN {annee} — {prop_nom.upper()}"
    c.font = Font(bold=True, size=16, color=BLANC, name="Arial")
    c.fill = PatternFill("solid", start_color=BLEU_DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:F2")
    ws["A2"].value = f"Généré le {date.today().strftime('%d/%m/%Y')}"
    ws["A2"].font = Font(italic=True, size=9, color="757575", name="Arial")
    ws["A2"].alignment = Alignment(horizontal="center")

    # KPIs
    kpis = [
        ("💰 CA Brut",        df["prix_brut"].sum()   if "prix_brut"    in df.columns else 0, "#,##0 €", BLEU_LIGHT),
        ("💵 CA Net",          df["prix_net"].sum()    if "prix_net"     in df.columns else 0, "#,##0 €", "C8E6C9"),
        ("🔖 Commissions",     df["commissions"].sum() if "commissions"  in df.columns else 0, "#,##0 €", "FFE0B2"),
        ("🧹 Ménage",          df["menage"].sum()      if "menage"       in df.columns else 0, "#,##0 €", "F3E5F5"),
        ("📅 Réservations",    len(df),                                                        "0",       BLEU_LIGHT),
        ("🌙 Nuits louées",    int(df["nuitees"].sum()) if "nuitees"     in df.columns else 0, "0",       "C8E6C9"),
    ]

    nuits = int(df["nuitees"].sum()) if "nuitees" in df.columns else 1
    ca_net = float(df["prix_net"].sum()) if "prix_net" in df.columns else 0
    revenu_nuit = round(ca_net / nuits) if nuits > 0 else 0
    kpis.append(("📈 Revenu / nuit", revenu_nuit, "#,##0 €", "FFF9C4"))

    nb_payes = int((df["paye"] == True).sum()) if "paye" in df.columns else 0
    taux_paiement = round(nb_payes / len(df) * 100, 1) if len(df) > 0 else 0
    kpis.append(("✅ Taux paiement", taux_paiement / 100, "0.0%", "C8E6C9"))

    row = 4
    _header_cell(ws, row, 1, "Indicateur", bg=BLEU_MED)
    _header_cell(ws, row, 2, "Valeur", bg=BLEU_MED)
    row += 1

    for label, val, fmt, bg in kpis:
        _data_cell(ws, row, 1, label, align="left", bg=GRIS_FOND, bold=True)
        _data_cell(ws, row, 2, val, fmt=fmt, bg=bg)
        ws.row_dimensions[row].height = 22
        row += 1

    # Répartition par plateforme
    row += 1
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"].value = "RÉPARTITION PAR PLATEFORME"
    ws[f"A{row}"].font = Font(bold=True, size=11, color=BLANC, name="Arial")
    ws[f"A{row}"].fill = PatternFill("solid", start_color=BLEU_MED)
    ws[f"A{row}"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[row].height = 25
    row += 1

    _header_cell(ws, row, 1, "Plateforme", bg=BLEU_LIGHT, fg="000000")
    _header_cell(ws, row, 2, "Réservations", bg=BLEU_LIGHT, fg="000000")
    _header_cell(ws, row, 3, "Nuits", bg=BLEU_LIGHT, fg="000000")
    _header_cell(ws, row, 4, "CA Net (€)", bg=BLEU_LIGHT, fg="000000")
    _header_cell(ws, row, 5, "% du CA", bg=BLEU_LIGHT, fg="000000")
    row += 1

    if "plateforme" in df.columns:
        plat = df.groupby("plateforme").agg(
            nb=("id", "count"),
            nuits=("nuitees", "sum"),
            ca_net=("prix_net", "sum"),
        ).reset_index()
        total_ca = plat["ca_net"].sum()

        for _, r in plat.iterrows():
            _data_cell(ws, row, 1, r["plateforme"], align="left")
            _data_cell(ws, row, 2, int(r["nb"]))
            _data_cell(ws, row, 3, int(r["nuits"]))
            _data_cell(ws, row, 4, float(r["ca_net"]), fmt="#,##0 €")
            _data_cell(ws, row, 5, float(r["ca_net"]) / total_ca if total_ca else 0, fmt="0.0%")
            ws.row_dimensions[row].height = 20
            row += 1

        # Totaux
        _data_cell(ws, row, 1, "TOTAL", align="left", bold=True, bg=BLEU_LIGHT)
        _data_cell(ws, row, 2, int(plat["nb"].sum()), bold=True, bg=BLEU_LIGHT)
        _data_cell(ws, row, 3, int(plat["nuits"].sum()), bold=True, bg=BLEU_LIGHT)
        _data_cell(ws, row, 4, float(plat["ca_net"].sum()), fmt="#,##0 €", bold=True, bg=BLEU_LIGHT)
        _data_cell(ws, row, 5, 1.0, fmt="0.0%", bold=True, bg=BLEU_LIGHT)

    # Largeurs colonnes
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12


# ── Onglet 2 : Détail mensuel ─────────────────────────────────────────────────

def _sheet_mensuel(wb, df, annee):
    ws = wb.create_sheet("📅 Mensuel")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:H1")
    ws["A1"].value = f"DÉTAIL MENSUEL {annee}"
    ws["A1"].font = Font(bold=True, size=14, color=BLANC, name="Arial")
    ws["A1"].fill = PatternFill("solid", start_color=BLEU_DARK)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    headers = ["Mois", "Réservations", "Nuits", "CA Brut (€)", "CA Net (€)",
               "Commissions (€)", "Ménage (€)", "Revenu/nuit (€)"]
    for col, h in enumerate(headers, 1):
        _header_cell(ws, 2, col, h, bg=BLEU_MED)

    # Préparer les données mensuelles
    MOIS = ["Jan","Fév","Mar","Avr","Mai","Jun","Jul","Aoû","Sep","Oct","Nov","Déc"]

    df["mois_num"] = df["date_arrivee"].dt.month
    monthly = df.groupby("mois_num").agg(
        nb=("id", "count"),
        nuits=("nuitees", "sum"),
        ca_brut=("prix_brut", "sum"),
        ca_net=("prix_net", "sum"),
        commissions=("commissions", "sum"),
        menage=("menage", "sum"),
    ).reset_index()

    data_rows = []
    for m in range(1, 13):
        row_data = monthly[monthly["mois_num"] == m]
        if not row_data.empty:
            r = row_data.iloc[0]
            rn = float(r["ca_net"]) / int(r["nuits"]) if int(r["nuits"]) > 0 else 0
            data_rows.append([MOIS[m-1], int(r["nb"]), int(r["nuits"]),
                              float(r["ca_brut"]), float(r["ca_net"]),
                              float(r["commissions"]), float(r["menage"]), rn])
        else:
            data_rows.append([MOIS[m-1], 0, 0, 0, 0, 0, 0, 0])

    for i, row_data in enumerate(data_rows):
        row_idx = i + 3
        bg = GRIS_FOND if i % 2 == 0 else BLANC
        _data_cell(ws, row_idx, 1, row_data[0], align="center", bold=True, bg=bg)
        _data_cell(ws, row_idx, 2, row_data[1], align="center", bg=bg)
        _data_cell(ws, row_idx, 3, row_data[2], align="center", bg=bg)
        for col in range(4, 9):
            fmt = "#,##0 €" if col != 8 else "#,##0 €"
            _data_cell(ws, row_idx, col, row_data[col-1], fmt=fmt, bg=bg)
        ws.row_dimensions[row_idx].height = 20

    # Totaux
    tot_row = 15
    _data_cell(ws, tot_row, 1, "TOTAL", align="center", bold=True, bg=BLEU_LIGHT)
    for col in range(2, 8):
        vals = [data_rows[m][col-1] for m in range(12)]
        _data_cell(ws, tot_row, col, sum(vals),
                   fmt="#,##0 €" if col > 3 else "0",
                   bold=True, bg=BLEU_LIGHT)
    # Revenu/nuit global
    tot_nuits = sum(r[2] for r in data_rows)
    tot_ca    = sum(r[4] for r in data_rows)
    _data_cell(ws, tot_row, 8, tot_ca / tot_nuits if tot_nuits else 0,
               fmt="#,##0 €", bold=True, bg=BLEU_LIGHT)

    # Graphique CA mensuel
    chart = BarChart()
    chart.type = "col"
    chart.title = f"CA Net mensuel {annee}"
    chart.y_axis.title = "€"
    chart.x_axis.title = "Mois"
    chart.style = 10
    chart.height = 12
    chart.width = 22

    data_ref = Reference(ws, min_col=5, min_row=2, max_row=14)
    cats_ref  = Reference(ws, min_col=1, min_row=3, max_row=14)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.series[0].graphicalProperties.solidFill = "3949AB"
    ws.add_chart(chart, "A17")

    for col, w in zip("ABCDEFGH", [10, 14, 10, 16, 16, 18, 14, 18]):
        ws.column_dimensions[get_column_letter(ord(col)-64)].width = w


# ── Onglet 3 : Liste réservations ─────────────────────────────────────────────

def _sheet_reservations(wb, df):
    ws = wb.create_sheet("📋 Réservations")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:J1")
    ws["A1"].value = "LISTE DES RÉSERVATIONS"
    ws["A1"].font = Font(bold=True, size=14, color=BLANC, name="Arial")
    ws["A1"].fill = PatternFill("solid", start_color=BLEU_DARK)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    headers = ["Client", "Plateforme", "Arrivée", "Départ", "Nuits",
               "CA Brut (€)", "Commissions (€)", "Ménage (€)", "CA Net (€)", "Payé"]
    for col, h in enumerate(headers, 1):
        _header_cell(ws, 2, col, h, bg=BLEU_MED)

    df_sorted = df.sort_values("date_arrivee")
    for i, (_, row) in enumerate(df_sorted.iterrows()):
        r = i + 3
        bg = GRIS_FOND if i % 2 == 0 else BLANC
        _data_cell(ws, r, 1,  str(row.get("nom_client", "")),   align="left",   bg=bg)
        _data_cell(ws, r, 2,  str(row.get("plateforme", "")),   align="center", bg=bg)
        _data_cell(ws, r, 3,  row["date_arrivee"].date() if pd.notna(row["date_arrivee"]) else "",
                   fmt="DD/MM/YYYY", align="center", bg=bg)
        _data_cell(ws, r, 4,  row["date_depart"].date()  if pd.notna(row["date_depart"])  else "",
                   fmt="DD/MM/YYYY", align="center", bg=bg)
        _data_cell(ws, r, 5,  int(row.get("nuitees", 0) or 0),  align="center", bg=bg)
        _data_cell(ws, r, 6,  float(row.get("prix_brut", 0) or 0),    fmt="#,##0 €", bg=bg)
        _data_cell(ws, r, 7,  float(row.get("commissions", 0) or 0),  fmt="#,##0 €", bg=bg)
        _data_cell(ws, r, 8,  float(row.get("menage", 0) or 0),       fmt="#,##0 €", bg=bg)
        _data_cell(ws, r, 9,  float(row.get("prix_net", 0) or 0),     fmt="#,##0 €", bg=bg)
        paye = row.get("paye", False)
        _data_cell(ws, r, 10, "✅" if paye else "⏳", align="center", bg=bg,
                   bold=True)
        ws.row_dimensions[r].height = 18

    for col, w in zip("ABCDEFGHIJ", [28, 12, 13, 13, 8, 14, 18, 12, 14, 8]):
        ws.column_dimensions[get_column_letter(ord(col)-64)].width = w


# ── Onglet 4 : Prévisions ─────────────────────────────────────────────────────

def _sheet_previsions(wb, df, prop_nom):
    ws = wb.create_sheet("🔮 Prévisions")
    ws.sheet_view.showGridLines = False

    today = pd.Timestamp(date.today())
    annee_courante = today.year

    ws.merge_cells("A1:F1")
    ws["A1"].value = "PRÉVISIONS — RÉSERVATIONS CONFIRMÉES"
    ws["A1"].font = Font(bold=True, size=14, color=BLANC, name="Arial")
    ws["A1"].fill = PatternFill("solid", start_color=BLEU_DARK)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Futures réservations (toutes années)
    df_future = df[df["date_arrivee"] > today].copy()
    df_future = df_future[df_future["plateforme"] != "Fermeture"]
    df_future["annee_arr"] = df_future["date_arrivee"].dt.year
    df_future["mois_num"]  = df_future["date_arrivee"].dt.month

    MOIS = ["Jan","Fév","Mar","Avr","Mai","Jun","Jul","Aoû","Sep","Oct","Nov","Déc"]

    row = 3
    for annee in sorted(df_future["annee_arr"].unique()):
        df_an = df_future[df_future["annee_arr"] == annee]

        ws.merge_cells(f"A{row}:F{row}")
        ws[f"A{row}"].value = f"RÉSERVATIONS FUTURES {annee}"
        ws[f"A{row}"].font = Font(bold=True, size=11, color=BLANC, name="Arial")
        ws[f"A{row}"].fill = PatternFill("solid", start_color=BLEU_MED)
        ws[f"A{row}"].alignment = Alignment(horizontal="center")
        ws.row_dimensions[row].height = 22
        row += 1

        _header_cell(ws, row, 1, "Mois",           bg=BLEU_LIGHT, fg="000000")
        _header_cell(ws, row, 2, "Réservations",   bg=BLEU_LIGHT, fg="000000")
        _header_cell(ws, row, 3, "Nuits",          bg=BLEU_LIGHT, fg="000000")
        _header_cell(ws, row, 4, "CA Net prévu (€)",bg=BLEU_LIGHT, fg="000000")
        _header_cell(ws, row, 5, "Payé (€)",       bg=BLEU_LIGHT, fg="000000")
        _header_cell(ws, row, 6, "En attente (€)", bg=BLEU_LIGHT, fg="000000")
        row += 1

        monthly = df_an.groupby("mois_num").agg(
            nb=("id", "count"),
            nuits=("nuitees", "sum"),
            ca_net=("prix_net", "sum"),
        ).reset_index()

        tot_nb = tot_nuits = tot_ca = tot_paye = tot_att = 0

        for m in range(1, 13):
            md = monthly[monthly["mois_num"] == m]
            if md.empty:
                continue
            r_data = md.iloc[0]
            df_mois = df_an[df_an["mois_num"] == m]
            paye    = float(df_mois[df_mois["paye"] == True]["prix_net"].sum())
            attente = float(df_mois[df_mois["paye"] != True]["prix_net"].sum())

            bg = GRIS_FOND if m % 2 == 0 else BLANC
            _data_cell(ws, row, 1, MOIS[m-1],          align="center", bold=True, bg=bg)
            _data_cell(ws, row, 2, int(r_data["nb"]),   align="center", bg=bg)
            _data_cell(ws, row, 3, int(r_data["nuits"]),align="center", bg=bg)
            _data_cell(ws, row, 4, float(r_data["ca_net"]), fmt="#,##0 €", bg=bg)
            _data_cell(ws, row, 5, paye,   fmt="#,##0 €", bg="C8E6C9")
            _data_cell(ws, row, 6, attente, fmt="#,##0 €", bg="FFECB3")
            ws.row_dimensions[row].height = 18

            tot_nb    += int(r_data["nb"])
            tot_nuits += int(r_data["nuits"])
            tot_ca    += float(r_data["ca_net"])
            tot_paye  += paye
            tot_att   += attente
            row += 1

        # Total
        _data_cell(ws, row, 1, f"TOTAL {annee}", align="center", bold=True, bg=BLEU_LIGHT)
        _data_cell(ws, row, 2, tot_nb,   align="center", bold=True, bg=BLEU_LIGHT)
        _data_cell(ws, row, 3, tot_nuits, align="center", bold=True, bg=BLEU_LIGHT)
        _data_cell(ws, row, 4, tot_ca,    fmt="#,##0 €", bold=True, bg=BLEU_LIGHT)
        _data_cell(ws, row, 5, tot_paye,  fmt="#,##0 €", bold=True, bg="A5D6A7")
        _data_cell(ws, row, 6, tot_att,   fmt="#,##0 €", bold=True, bg="FFE082")
        ws.row_dimensions[row].height = 22
        row += 2

    for col, w in zip("ABCDEF", [12, 16, 10, 18, 16, 16]):
        ws.column_dimensions[get_column_letter(ord(col)-64)].width = w
